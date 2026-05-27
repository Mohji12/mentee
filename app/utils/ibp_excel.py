"""
Excel generation utilities for IBP form responses
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from app.utils.ibp_questions import IBP_QUESTIONS, IBP_OPTIONS


def generate_ibp_excel(
    student_usn: str,
    student_name: str,
    student_program: str,
    responses: dict,
    submitted_at: str,
) -> BytesIO:
    """
    Generate Excel file for a student's IBP responses.

    Args:
        student_usn: Student USN
        student_name: Student name
        student_program: Student program
        responses: Dictionary of question_number -> answer (e.g., {1: 3, 2: 5, ...} or {"1": "3", ...})
        submitted_at: Submission datetime string

    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Student Information
    ws_info = wb.create_sheet("Student Information")
    ws_info.append(["Student USN", student_usn])
    ws_info.append(["Student Name", student_name])
    ws_info.append(["Program", student_program])
    ws_info.append(["Submitted At", submitted_at])
    ws_info.append([])

    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    # Sheet 2: Responses — Question and Answer for each statement
    ws_responses = wb.create_sheet("IBP Responses")

    # Headers: Question (statement) and Given Answer
    headers = [
        "Sl #",
        "Question",
        "Given Answer",
    ]
    ws_responses.append(headers)

    # Style headers
    for cell in ws_responses[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    # Normalize responses to integer keys and values
    normalized_responses = {}
    for key, value in responses.items():
        try:
            key_int = int(key) if isinstance(key, str) else key
            val_int = int(value) if isinstance(value, str) else value
            normalized_responses[key_int] = val_int
        except (ValueError, TypeError) as e:
            # Skip invalid entries but log for debugging
            print(f"Warning: Skipping invalid response entry: key={key}, value={value}, error={e}")
            continue

    # Debug: Check if IBP_QUESTIONS is loaded
    if not IBP_QUESTIONS:
        raise ValueError("IBP_QUESTIONS dictionary is empty. Cannot generate Excel.")

    # Debug: Check if responses are present
    if not normalized_responses:
        print(f"Warning: No valid responses found in responses dict: {responses}")

    # Iterate through all questions and add responses
    rows_added = 0
    for q_num in sorted(IBP_QUESTIONS.keys()):
        if q_num not in IBP_QUESTIONS:
            print(f"Warning: Question {q_num} not found in IBP_QUESTIONS")
            continue
            
        question_data = IBP_QUESTIONS[q_num]
        if not isinstance(question_data, dict) or "text" not in question_data:
            print(f"Warning: Invalid question data for question {q_num}: {question_data}")
            continue
            
        answer = normalized_responses.get(q_num)
        answer_str = str(answer) if answer is not None else ""
        answer_label = IBP_OPTIONS.get(answer_str, "") if answer_str else ""
        # Show answer as both number and label, e.g. "3 - Sometimes"
        given_answer = f"{answer_str} - {answer_label}" if answer_str and answer_label else (answer_label or answer_str or "Not answered")

        row = [
            q_num,
            question_data["text"],
            given_answer,
        ]
        ws_responses.append(row)
        rows_added += 1
    
    # Validate that rows were added
    if rows_added == 0:
        raise ValueError(f"No rows were added to Excel. IBP_QUESTIONS has {len(IBP_QUESTIONS)} questions, normalized_responses has {len(normalized_responses)} responses.")
    
    print(f"Successfully added {rows_added} rows to IBP Excel for student {student_usn}")

    # Style data rows
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws_responses.iter_rows(min_row=2, max_row=ws_responses.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column == 1:  # Sl #
                cell.alignment = Alignment(horizontal="center")
            elif cell.column == 3:  # Given Answer
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
                if cell.value:
                    cell.font = Font(bold=True, color="006100")

    # Column widths
    ws_responses.column_dimensions["A"].width = 8
    ws_responses.column_dimensions["B"].width = 80
    ws_responses.column_dimensions["C"].width = 24

    ws_responses.freeze_panes = "A2"

    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 40

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
