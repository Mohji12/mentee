"""
Excel generation utilities for 16PF form responses
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from app.utils.pf16_questions import PF16_QUESTIONS


def generate_pf16_excel(student_usn: str, student_name: str, student_program: str, responses: dict, submitted_at: str) -> BytesIO:
    """
    Generate Excel file for a student's 16PF responses.
    
    Args:
        student_usn: Student USN
        student_name: Student name
        student_program: Student program
        responses: Dictionary of question_number -> answer (e.g., {1: "a", 2: "b", ...})
        submitted_at: Submission datetime string
    
    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: Student Information
    ws_info = wb.create_sheet("Student Information")
    ws_info.append(["Student USN", student_usn])
    ws_info.append(["Student Name", student_name])
    ws_info.append(["Program", student_program])
    ws_info.append(["Submitted At", submitted_at])
    ws_info.append([])
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Sheet 2: Responses
    ws_responses = wb.create_sheet("16PF Responses")
    
    # Headers
    headers = ["Question #", "Question Text", "Answer Selected", "Option A", "Option B", "Option C"]
    ws_responses.append(headers)
    
    # Style headers
    for cell in ws_responses[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Add data rows - iterate through all questions (1-185) to ensure all are shown
    # Normalize responses dictionary to use integer keys
    normalized_responses = {}
    for key, value in responses.items():
        try:
            key_int = int(key) if isinstance(key, str) else key
            normalized_responses[key_int] = value
        except (ValueError, TypeError) as e:
            # Skip invalid entries but log for debugging
            print(f"Warning: Skipping invalid response entry: key={key}, value={value}, error={e}")
            continue

    # Debug: Check if PF16_QUESTIONS is loaded
    if not PF16_QUESTIONS:
        raise ValueError("PF16_QUESTIONS dictionary is empty. Cannot generate Excel.")

    # Debug: Check if responses are present
    if not normalized_responses:
        print(f"Warning: No valid responses found in responses dict: {responses}")

    # Iterate through all questions in PF16_QUESTIONS (1-185)
    rows_added = 0
    for q_num in sorted(PF16_QUESTIONS.keys()):
        if q_num not in PF16_QUESTIONS:
            print(f"Warning: Question {q_num} not found in PF16_QUESTIONS")
            continue
            
        question_data = PF16_QUESTIONS[q_num]
        if not isinstance(question_data, dict) or "text" not in question_data:
            print(f"Warning: Invalid question data for question {q_num}: {question_data}")
            continue
        
        answer = normalized_responses.get(q_num, "")
        answer_str = answer.upper() if answer else ""
        
        # Ensure options exist
        options = question_data.get("options", {})
        
        row = [
            q_num,
            question_data["text"],
            answer_str,
            options.get("a", ""),
            options.get("b", ""),
            options.get("c", "")
        ]
        ws_responses.append(row)
        rows_added += 1
    
    # Validate that rows were added
    if rows_added == 0:
        raise ValueError(f"No rows were added to Excel. PF16_QUESTIONS has {len(PF16_QUESTIONS)} questions, normalized_responses has {len(normalized_responses)} responses.")
    
    print(f"Successfully added {rows_added} rows to 16PF Excel for student {student_usn}")
    
    # Style data rows
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for row in ws_responses.iter_rows(min_row=2, max_row=ws_responses.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column == 1:  # Question #
                cell.alignment = Alignment(horizontal="center")
            elif cell.column == 3:  # Answer Selected
                cell.alignment = Alignment(horizontal="center")
                if cell.value:
                    cell.font = Font(bold=True, color="006100")
    
    # Auto-adjust column widths
    ws_responses.column_dimensions['A'].width = 12  # Question #
    ws_responses.column_dimensions['B'].width = 80  # Question Text
    ws_responses.column_dimensions['C'].width = 15  # Answer Selected
    ws_responses.column_dimensions['D'].width = 40  # Option A
    ws_responses.column_dimensions['E'].width = 10  # Option B
    ws_responses.column_dimensions['F'].width = 40  # Option C
    
    # Freeze header row
    ws_responses.freeze_panes = "A2"
    
    # Auto-adjust info sheet
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 40
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
